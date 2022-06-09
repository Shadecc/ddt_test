# coding=utf8
import time
import unittest
import warnings
import openpyxl as openpyxl
from ddt import ddt, data, unpack
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.common.by import By

from ddt_test.slide_check import slide

file_name = 'data.xlsx'


# 读取excel里面的数据xlrd, openpyxl
def read_excel():
    """
    读取excel
    :return:
    """
    # 加载excel表格
    xls = openpyxl.load_workbook(file_name)
    sheet1 = xls['Sheet1']
    all_list = []
    for row in range(2, sheet1.max_row + 1):
        row_list = []
        for col in range(1, sheet1.max_column + 1):
            row_list.append(sheet1.cell(row, col).value)
        all_list.append(row_list)
    return all_list


@ddt
class HtmlTest(unittest.TestCase):
    @classmethod
    def setUpClass(cls) -> None:
        options = webdriver.FirefoxOptions()
        options.add_argument('--disable-blink-features=AutomationControlled')
        cls.driver = webdriver.Firefox(options=options)
        cls.driver.get('https://www.douyin.com/')
        warnings.simplefilter("ignore", ResourceWarning)

    @data(*read_excel())  # 分离数据
    @unpack
    def test_01_login(self, flag, content):
        print(flag, content)
        # # 清空输入框
        # self.driver.find_element(By.CLASS_NAME, 'KLzwyB7s mV31vsEW').clear()
        # 在输入框输入内容
        self.driver.find_element(By.CLASS_NAME, 'KLzwyB7s').send_keys(content)
        time.sleep(1)

        self.driver.find_element(By.CLASS_NAME, 'KLzwyB7s').send_keys(content).clear()
        # # 封装一个函数，用来判断属性值是否存在
        # def test_exceptions(xpath):
        #     try:
        #         self.driver.find_element_by_xpath(xpath)
        #         return True
        #     except:
        #         return False
        # # 判断
        # if test_exceptions('/html/body/div[3]/div/div[1]'):
        #     # 停一下，等待出现
        #     time.sleep(2)
        #     # 滑动验证码
        #     slide(self.driver)
        #     time.sleep(5)
        #     self.driver.switch_to.default_content()
        #
        # # 点击搜索
        # self.driver.find_element(By.CLASS_NAME, 'kT0ePbyQ').click()
        # time.sleep(3)
        # # 找到搜索结果
        # num = self.driver.find_element(By.CLASS_NAME, 'EO7og0qV').text
        # time.sleep(30)
        #
        # """
        # 写入数据
        # """
        # # 加载已存在的wookbook对象
        # wb = load_workbook(file_name)
        # wb1 = wb.active  # 激活sheet
        # # 根据实际aim_row行，aim_col列修改数据即可
        # aim_row = flag + 1
        # aim_col = 3
        # # 往sheet中的第flag + 1行第三列写入num的数据
        # wb1.cell(aim_row, aim_col, num)
        # # 保存
        # wb.save(file_name)

    @classmethod
    def tearDownClass(cls) -> None:
        cls.driver.close()


if __name__ == '__main__':
    read_excel()